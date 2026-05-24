"""
Generadores de archivos: Excel (openpyxl) y PDF (reportlab).

Cada función retorna (bytes, filename, mime_type) para que pueda servirse
igual desde:
    REST  →  HttpResponse(file_bytes, content_type=mime)
    gRPC  →  reportes_pb2.FileResponse(file_bytes=..., file_name=..., mime_type=...)

Los datos se obtienen vía gRPC (módulo grpc_clients).
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from . import grpc_clients

AZUL_BUAP  = "00637C"
GRIS_CLARO = "EAEAEA"
NEGRO      = "000000"

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_MIME  = "application/pdf"


# ─── Helpers de estilo Excel ──────────────────────────────────────────────────
def _apply_header_style(cell):
    cell.font      = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill      = PatternFill("solid", fgColor=AZUL_BUAP)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin           = Side(border_style="thin", color=NEGRO)
    cell.border    = Border(top=thin, left=thin, right=thin, bottom=thin)


def _apply_cell_style(cell, align="left"):
    cell.font      = Font(name="Calibri", size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    thin           = Side(border_style="thin", color="999999")
    cell.border    = Border(top=thin, left=thin, right=thin, bottom=thin)


def _info_block(ws, row_start: int, materia: dict) -> int:
    """Escribe el bloque institucional + datos de la materia. Devuelve la siguiente fila libre."""
    ws.cell(row=row_start,     column=1, value="BENEMÉRITA UNIVERSIDAD AUTÓNOMA DE PUEBLA").font = Font(bold=True, size=12)
    ws.cell(row=row_start + 1, column=1, value="Facultad de Ciencias de la Computación").font    = Font(bold=True, size=11)
    ws.cell(row=row_start + 2, column=1, value=f"Generado: {datetime.now():%Y-%m-%d %H:%M}").font = Font(italic=True, size=9)

    info = [
        ("NRC:",     materia.get("nrc",       "")),
        ("Clave:",   materia.get("clave",     "")),
        ("Materia:", materia.get("nombre",    "")),
        ("Sección:", materia.get("seccion",   "")),
        ("Periodo:", materia.get("periodo_id","")),
        ("Horario:", materia.get("horario",   "")),
    ]
    r = row_start + 4
    for label, value in info:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=str(value))
        r += 1
    return r + 1


# ─── CALIFICACIONES — Excel ───────────────────────────────────────────────────
def generar_calificaciones_xlsx(materia_id) -> tuple[bytes, str, str]:
    materia = grpc_clients.get_materia(materia_id)
    alumnos = grpc_clients.get_concentrado(materia_id)
    stats   = grpc_clients.get_estadisticas_materia(materia_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Calificaciones"

    for i, w in enumerate([6, 16, 42, 14, 12, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    next_row = _info_block(ws, row_start=1, materia=materia)

    headers = ["#", "Matrícula", "Nombre completo", "Promedio real", "Redondeado", "Estado"]
    for col, h in enumerate(headers, start=1):
        _apply_header_style(ws.cell(row=next_row, column=col, value=h))
    ws.row_dimensions[next_row].height = 28

    body_row = next_row + 1
    for idx, alu in enumerate(alumnos, start=1):
        row_data = [
            idx,
            alu.get("matricula",           ""),
            alu.get("nombre_completo",      ""),
            alu.get("promedio_real",         0),
            alu.get("promedio_redondeado",   0),
            alu.get("estado",               ""),
        ]
        for col, v in enumerate(row_data, start=1):
            c = ws.cell(row=body_row, column=col, value=v)
            _apply_cell_style(c, align="center" if col in (1, 4, 5, 6) else "left")
            if col == 6:
                if v == "APROBADO":
                    c.fill = PatternFill("solid", fgColor="C6EFCE")
                elif v == "REPROBADO":
                    c.fill = PatternFill("solid", fgColor="FFC7CE")
        body_row += 1

    body_row += 1
    ws.cell(row=body_row, column=1, value="Resumen del grupo").font = Font(bold=True, size=11)
    body_row += 1
    for label, val in [
        ("Total de alumnos:",   stats["total_alumnos"]),
        ("Aprobados:",          stats["aprobados"]),
        ("Reprobados:",         stats["reprobados"]),
        ("Promedio grupal:",    stats["promedio_grupal"]),
        ("Calificación máxima:", stats["calificacion_maxima"]),
        ("Calificación mínima:", stats["calificacion_minima"]),
    ]:
        ws.cell(row=body_row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=body_row, column=2, value=val)
        body_row += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), f"calificaciones_{materia.get('nrc') or materia_id}.xlsx", XLSX_MIME


# ─── CALIFICACIONES — PDF ─────────────────────────────────────────────────────
def generar_calificaciones_pdf(materia_id) -> tuple[bytes, str, str]:
    materia = grpc_clients.get_materia(materia_id)
    alumnos = grpc_clients.get_concentrado(materia_id)
    stats   = grpc_clients.get_estadisticas_materia(materia_id)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm,
        title=f"Calificaciones {materia.get('nombre', '')}",
    )

    styles   = getSampleStyleSheet()
    title_st = ParagraphStyle("title", parent=styles["Title"],  fontSize=14,
                               textColor=colors.HexColor(f"#{AZUL_BUAP}"), alignment=1, spaceAfter=4)
    sub_st   = ParagraphStyle("sub",   parent=styles["Normal"], fontSize=10, alignment=1, spaceAfter=12)
    label_st = ParagraphStyle("label", parent=styles["Normal"], fontSize=9,  leading=12)

    story = [
        Paragraph("BENEMÉRITA UNIVERSIDAD AUTÓNOMA DE PUEBLA", title_st),
        Paragraph("Facultad de Ciencias de la Computación", sub_st),
        Paragraph(f"<b>Reporte de Calificaciones</b> &nbsp;&nbsp; Generado: {datetime.now():%Y-%m-%d %H:%M}", sub_st),
    ]

    info_data = [
        ["NRC:",     materia.get("nrc",       ""), "Clave:",   materia.get("clave",   "")],
        ["Materia:", materia.get("nombre",    ""), "Sección:", materia.get("seccion", "")],
        ["Periodo:", materia.get("periodo_id",""), "Estado:",  materia.get("estado",  "")],
        ["Horario:", materia.get("horario",   ""), "",         ""],
    ]
    info_tbl = Table(info_data, colWidths=[2.2*cm, 6.5*cm, 2.2*cm, 6.5*cm])
    info_tbl.setStyle(TableStyle([
        ("FONTNAME",  (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE",  (0, 0), (-1, -1), 9),
        ("FONTNAME",  (0, 0), (0, -1),  "Helvetica-Bold"),
        ("FONTNAME",  (2, 0), (2, -1),  "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story += [info_tbl, Spacer(1, 0.5*cm)]

    headers = ["#", "Matrícula", "Nombre completo", "Prom. real", "Redond.", "Estado"]
    rows = [headers] + [
        [
            str(i),
            a.get("matricula",          ""),
            a.get("nombre_completo",     ""),
            f"{a.get('promedio_real', 0):.2f}",
            str(a.get("promedio_redondeado", 0)),
            a.get("estado", ""),
        ]
        for i, a in enumerate(alumnos, start=1)
    ]

    tbl = Table(rows, colWidths=[0.8*cm, 2.4*cm, 7.0*cm, 2.0*cm, 1.6*cm, 2.4*cm], repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{AZUL_BUAP}")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1,-1), 9),
        ("ALIGN",      (0, 0), (-1,-1), "CENTER"),
        ("ALIGN",      (2, 1), (2, -1), "LEFT"),
        ("VALIGN",     (0, 0), (-1,-1), "MIDDLE"),
        ("GRID",       (0, 0), (-1,-1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1,-1), [colors.white, colors.HexColor("#F4F7F9")]),
    ]
    for i, a in enumerate(alumnos, start=1):
        if a.get("estado") == "APROBADO":
            style_cmds.append(("BACKGROUND", (5, i), (5, i), colors.HexColor("#C6EFCE")))
        elif a.get("estado") == "REPROBADO":
            style_cmds.append(("BACKGROUND", (5, i), (5, i), colors.HexColor("#FFC7CE")))
    tbl.setStyle(TableStyle(style_cmds))
    story += [tbl, Spacer(1, 0.6*cm)]

    resumen_data = [
        ["Total alumnos:", str(stats["total_alumnos"]),   "Promedio grupal:",  f"{stats['promedio_grupal']:.2f}"],
        ["Aprobados:",     str(stats["aprobados"]),       "Calif. máxima:",    f"{stats['calificacion_maxima']:.2f}"],
        ["Reprobados:",    str(stats["reprobados"]),      "Calif. mínima:",    f"{stats['calificacion_minima']:.2f}"],
    ]
    resumen_tbl = Table(resumen_data, colWidths=[3.5*cm, 2.5*cm, 3.5*cm, 2.5*cm])
    resumen_tbl.setStyle(TableStyle([
        ("FONTNAME",  (0, 0), (-1,-1), "Helvetica"),
        ("FONTSIZE",  (0, 0), (-1,-1), 9),
        ("FONTNAME",  (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME",  (2, 0), (2, -1), "Helvetica-Bold"),
        ("BOX",       (0, 0), (-1,-1), 0.4, colors.grey),
        ("INNERGRID", (0, 0), (-1,-1), 0.2, colors.lightgrey),
        ("BACKGROUND",(0, 0), (-1,-1), colors.HexColor(f"#{GRIS_CLARO}")),
        ("LEFTPADDING",  (0, 0), (-1,-1), 6),
        ("RIGHTPADDING", (0, 0), (-1,-1), 6),
    ]))
    story += [Paragraph("<b>Resumen del grupo</b>", label_st), Spacer(1, 0.15*cm), resumen_tbl]

    doc.build(story)
    return buf.getvalue(), f"calificaciones_{materia.get('nrc') or materia_id}.pdf", PDF_MIME


# ─── ASISTENCIAS — Excel ──────────────────────────────────────────────────────
def generar_asistencias_xlsx(materia_id) -> tuple[bytes, str, str]:
    materia = grpc_clients.get_materia(materia_id)
    alumnos = grpc_clients.get_alumnos_by_materia(materia_id)
    stats   = grpc_clients.get_estadisticas_asistencia(materia_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencias"

    for i, w in enumerate([6, 16, 42, 14, 14, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    next_row = _info_block(ws, row_start=1, materia=materia)

    headers = ["#", "Matrícula", "Nombre completo", "Presentes", "Retardos", "Ausencias", "% Asist."]
    for col, h in enumerate(headers, start=1):
        _apply_header_style(ws.cell(row=next_row, column=col, value=h))
    ws.row_dimensions[next_row].height = 28

    total_sesiones = stats.get("total_sesiones", 0) or 1
    body_row = next_row + 1

    for idx, alu in enumerate(alumnos, start=1):
        regs = grpc_clients.get_asistencia_alumno(alu["alumno_id"], materia_id)
        pres = sum(1 for r in regs if r["estado"] == "PRESENTE")
        ret  = sum(1 for r in regs if r["estado"] == "RETARDO")
        aus  = sum(1 for r in regs if r["estado"] == "AUSENTE")
        if aus == 0 and total_sesiones > (pres + ret):
            aus = total_sesiones - pres - ret
        pct = round((pres + ret) / total_sesiones * 100, 1)

        for col, v in enumerate([idx, alu["matricula"], alu["nombre_completo"],
                                  pres, ret, aus, f"{pct}%"], start=1):
            c = ws.cell(row=body_row, column=col, value=v)
            _apply_cell_style(c, align="left" if col == 3 else "center")
        body_row += 1

    body_row += 1
    ws.cell(row=body_row, column=1, value="Resumen del grupo").font = Font(bold=True, size=11)
    body_row += 1
    for label, val in [
        ("Total sesiones:",     stats.get("total_sesiones",    0)),
        ("Total asistencias:",  stats.get("total_asistencias", 0)),
        ("Total retardos:",     stats.get("total_retardos",    0)),
        ("Total ausencias:",    stats.get("total_ausencias",   0)),
        ("% asistencia grupal:", f"{stats.get('porcentaje_asistencia', 0)}%"),
    ]:
        ws.cell(row=body_row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=body_row, column=2, value=val)
        body_row += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), f"asistencias_{materia.get('nrc') or materia_id}.xlsx", XLSX_MIME


# ─── ASISTENCIAS — PDF ────────────────────────────────────────────────────────
def generar_asistencias_pdf(materia_id) -> tuple[bytes, str, str]:
    materia = grpc_clients.get_materia(materia_id)
    alumnos = grpc_clients.get_alumnos_by_materia(materia_id)
    stats   = grpc_clients.get_estadisticas_asistencia(materia_id)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm,
        title=f"Asistencias {materia.get('nombre', '')}",
    )

    styles   = getSampleStyleSheet()
    title_st = ParagraphStyle("title", parent=styles["Title"],  fontSize=14,
                               textColor=colors.HexColor(f"#{AZUL_BUAP}"), alignment=1, spaceAfter=4)
    sub_st   = ParagraphStyle("sub",   parent=styles["Normal"], fontSize=10, alignment=1, spaceAfter=12)
    label_st = ParagraphStyle("label", parent=styles["Normal"], fontSize=9,  leading=12)

    story = [
        Paragraph("BENEMÉRITA UNIVERSIDAD AUTÓNOMA DE PUEBLA", title_st),
        Paragraph("Facultad de Ciencias de la Computación", sub_st),
        Paragraph(f"<b>Concentrado de Asistencias</b> &nbsp;&nbsp; Generado: {datetime.now():%Y-%m-%d %H:%M}", sub_st),
    ]

    info_data = [
        ["NRC:",     materia.get("nrc",       ""), "Clave:",    materia.get("clave",   "")],
        ["Materia:", materia.get("nombre",    ""), "Sección:",  materia.get("seccion", "")],
        ["Periodo:", materia.get("periodo_id",""), "Sesiones:", str(stats.get("total_sesiones", 0))],
    ]
    info_tbl = Table(info_data, colWidths=[2.2*cm, 6.5*cm, 2.2*cm, 6.5*cm])
    info_tbl.setStyle(TableStyle([
        ("FONTNAME",  (0, 0), (-1,-1), "Helvetica"),
        ("FONTSIZE",  (0, 0), (-1,-1), 9),
        ("FONTNAME",  (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME",  (2, 0), (2, -1), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1,-1), 4),
    ]))
    story += [info_tbl, Spacer(1, 0.5*cm)]

    total_sesiones = stats.get("total_sesiones", 0) or 1
    headers = ["#", "Matrícula", "Nombre completo", "Pres.", "Ret.", "Aus.", "% Asist."]
    rows = [headers]
    for i, alu in enumerate(alumnos, start=1):
        regs = grpc_clients.get_asistencia_alumno(alu["alumno_id"], materia_id)
        pres = sum(1 for r in regs if r["estado"] == "PRESENTE")
        ret  = sum(1 for r in regs if r["estado"] == "RETARDO")
        aus  = sum(1 for r in regs if r["estado"] == "AUSENTE")
        if aus == 0 and total_sesiones > (pres + ret):
            aus = total_sesiones - pres - ret
        pct = round((pres + ret) / total_sesiones * 100, 1)
        rows.append([str(i), alu["matricula"], alu["nombre_completo"],
                     str(pres), str(ret), str(aus), f"{pct}%"])

    tbl = Table(rows, colWidths=[0.8*cm, 2.4*cm, 6.4*cm, 1.6*cm, 1.6*cm, 1.6*cm, 1.8*cm], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{AZUL_BUAP}")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1,-1), 9),
        ("ALIGN",      (0, 0), (-1,-1), "CENTER"),
        ("ALIGN",      (2, 1), (2, -1), "LEFT"),
        ("VALIGN",     (0, 0), (-1,-1), "MIDDLE"),
        ("GRID",       (0, 0), (-1,-1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1,-1), [colors.white, colors.HexColor("#F4F7F9")]),
    ]))
    story += [tbl, Spacer(1, 0.6*cm)]

    resumen_data = [
        ["Total sesiones:",  str(stats.get("total_sesiones",    0)), "% Asist. grupal:", f"{stats.get('porcentaje_asistencia', 0)}%"],
        ["Asistencias:",     str(stats.get("total_asistencias", 0)), "Retardos:",        str(stats.get("total_retardos", 0))],
        ["Ausencias:",       str(stats.get("total_ausencias",   0)), "",                 ""],
    ]
    resumen_tbl = Table(resumen_data, colWidths=[3.5*cm, 2.5*cm, 3.5*cm, 2.5*cm])
    resumen_tbl.setStyle(TableStyle([
        ("FONTNAME",  (0, 0), (-1,-1), "Helvetica"),
        ("FONTSIZE",  (0, 0), (-1,-1), 9),
        ("FONTNAME",  (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME",  (2, 0), (2, -1), "Helvetica-Bold"),
        ("BOX",       (0, 0), (-1,-1), 0.4, colors.grey),
        ("INNERGRID", (0, 0), (-1,-1), 0.2, colors.lightgrey),
        ("BACKGROUND",(0, 0), (-1,-1), colors.HexColor(f"#{GRIS_CLARO}")),
        ("LEFTPADDING",  (0, 0), (-1,-1), 6),
        ("RIGHTPADDING", (0, 0), (-1,-1), 6),
    ]))
    story += [Paragraph("<b>Resumen del grupo</b>", label_st), Spacer(1, 0.15*cm), resumen_tbl]

    doc.build(story)
    return buf.getvalue(), f"asistencias_{materia.get('nrc') or materia_id}.pdf", PDF_MIME


# ─── Registro de generadores ──────────────────────────────────────────────────
REGISTRO = {
    ("CALIFICACIONES", "XLSX"): generar_calificaciones_xlsx,
    ("CALIFICACIONES", "XLS"):  generar_calificaciones_xlsx,
    ("CALIFICACIONES", "PDF"):  generar_calificaciones_pdf,
    ("ASISTENCIAS",    "XLSX"): generar_asistencias_xlsx,
    ("ASISTENCIAS",    "XLS"):  generar_asistencias_xlsx,
    ("ASISTENCIAS",    "PDF"):  generar_asistencias_pdf,
}